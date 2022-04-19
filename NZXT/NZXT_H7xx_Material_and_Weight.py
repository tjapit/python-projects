import requests
import lxml.html as lh
import openpyxl
import time

### Initialize url, NZXT product page
url = 'https://www.nzxt.com/products/'

# Get page with requests.get()
page = requests.get(url)
print(page)     # Response code, should be 200

# Get html content with lh.fromstring()
doc = lh.fromstring(page.content)

# Get table rows from product page with xpath() by specifying tag, in this case tr
prod_name = doc.xpath('//h3//a')           # Product names xpath

### Get products with H7xx in their name
product_h7xx = []
i = 0

# NOTE: xpath extracts in the form of a list
for t in prod_name:

    # Get text from //h3//a// , (...)[0] is to extract it from the list
    name = t.xpath('text()')[0]

    # Get link from //h3//a// , NOTE: don't need '//' once in the path
    href = t.xpath('@href')[0]

    if 'H7' in name:
        product_h7xx.append((name, href, []))

print(product_h7xx)      # Checking list

### Get rid of /products in initial url

# Find index up to the beginning of /products and get rid of it
ext = url.index('/p')
home_url = url[:ext]

# Write name and weight of the product to an excel sheet

# Define initial column in excel sheet
name_column = 1
materials_column = 2
weight_column = 3
j = 2       # Start from second row, first row is header (OPENPYXL IS 1-INDEXING)
x = 0       # Loop indexing

# Headers
headers = ['Case', 'Material(s)', 'Weight']

# Creating new workbook and inserting headers to first row
wb = openpyxl.Workbook()
ws = wb.active
filename = 'F:/Python Projects/Scraping/NZXT_H7xx_Material_and_Weight.xlsx'

# Writing headers to first rows of the excel sheet
for h in range(len(headers)):
    ws.cell(row=1, column=h+1).value =  headers[h]

# Looping through the tuples in the h700 line-up
for t in product_h7xx:

    # Appending the url of the specific h7xx series
    url = home_url + t[1]

    # Requesting the page from the specific product
    page = requests.get(url)

    # Checking status code, 200 = gtg
    print(page)

    # Parsing html
    content = lh.fromstring(page.content)

    # Saving the table rows to tr_elements (ones with <tr> tags within the table)
    tr_elements = content.xpath('//tr')

    # Processing each row in tr_elements
    for r in tr_elements:
        # try except clause to catch empty lists from causing an error
        try:

            # Navigating html 'directory' within <tr> exists two <td> tags with "name" and "value" classes
            # appended ...[0] at the end because ...//text() pulls out a list
            table_header = r.xpath('td[@class="name"]//text()')[0]
            table_value = r.xpath('td[@class="value"]//text()')[0]

            # Checking extracted data (not needed really, more for aesthetics and to calm my anxious mind)
            print(table_header + ' : ' + table_value)

            # Appending Materials and Weight data to the list within the product_h7xx tuples
            if table_header == 'Material(s)' or table_header == 'Weight':
                product_h7xx[x][2].append(table_value)
        except:
            pass

    # Name of product to the first column of the j'th row
    ws.cell(row=j, column=name_column).value = str(product_h7xx[x][0])        #class 'lxml.etree.ElementUnicodeResult'

    # Material of product to second column of the j'th row
    ws.cell(row=j, column=materials_column).value = str(product_h7xx[x][2][0])     #class 'lxml.etree.ElementUnicodeResult'

    # catching the last row because it doesn't have a FUCKING WEIGHT INFORMATION AND CAUSES ERROR TO THE WHOLE CODE
    try:
        # Weight of product to third column of the j'th row
        ws.cell(row=j, column=weight_column).value = str(product_h7xx[x][2][1])        #class 'lxml.etree.ElementUnicodeResult'
    except:
        pass

    j+=1
    x+=1



# Saving changes
wb.save(filename)


