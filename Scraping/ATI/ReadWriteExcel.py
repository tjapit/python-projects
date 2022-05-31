import openpyxl
# Reading from Excel
path = 'I:\Timothy J\Rotation02\Routing Times for Configurations.xlsx'
wb = openpyxl.load_workbook(path)   # Loading workbook
sheet_name = '310 Configs. '        # Specifying sheet name
active_sheet = wb[sheet_name]              # Active sheet
itemID = active_sheet.cell(row=2, column=1)    # Indexing start from 1 instead of 0
print(itemID.value)                 # xxx.cell only gives indexing, xxx.value gives the actual value of the cell

# Writing to Excel
path = 'I:\Timothy J\Rotation02\Routing Times for Configurations.xlsx'
wb = openpyxl.load_workbook(path)
sheet_name = '310 Configs. '
active_sheet = wb[sheet_name]
active_sheet['D10'].value = 'You succ XD'      # Writing into specific row and column
wb.save(path)       # Need to save the file after any changes made, otherwise none will show


